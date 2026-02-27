from django.shortcuts import render,redirect,get_object_or_404
import json
import openpyxl
from django.utils import timezone
from django.http import HttpResponse
from .models import *
from Adminapp.models import *
from django.db.models import Q
from django.contrib.auth.decorators import login_required
from django.http import JsonResponse
from django.db.models import Exists, OuterRef
from openpyxl.styles import Font
from openpyxl import Workbook
import re
from django.core.paginator import Paginator
from django.contrib import messages

def add_consignee_ajax(request):
    if request.method == "POST":
        name = request.POST.get("billing_consignee_name")

        consignee = BillingConsignee.objects.create(
            billing_consignee_name=name,
        )

        return JsonResponse({
            "success": True,
            "id": consignee.billing_consignee_id,
            "name": consignee.billing_consignee_name
        })

    return JsonResponse({"success": False})

def accounts_dashboard(req):
    return render(req,"accounts/accounts_dashboard.html")

def billing_consignor_manage(req,billing_consignor_id=None):
    query = req.GET.get('q', '').strip()

    shippers = BillingConsignor.objects.filter(billing_consignor_active=True).order_by('-billing_consignor_id')

    if query:
        shippers = shippers.filter(
            Q(billing_consignor_name__icontains=query) |
            Q(billing_consignor_phone__icontains=query) |
            Q(billing_consignor_gst__icontains=query) |
            Q(billing_consingor_address__icontains=query)
        )
    instance = None
    mode = req.GET.get('mode', 'list')
    if billing_consignor_id:
        mode = 'edit'
        instance = get_object_or_404(BillingConsignor, billing_consignor_id=billing_consignor_id)
    if req.method == 'POST':
        con = instance if billing_consignor_id else BillingConsignor()

        con.billing_consignor_name = req.POST.get('consignor_name', '').strip()
        con.billing_consignor_phone = req.POST.get('consignor_phone', '').strip()
        con.billing_consignor_gst = req.POST.get('gst_no', '').strip()
        con.billing_consignor_gsttype = req.POST.get('gst_type', '').strip()
        con.billing_consingor_address = req.POST.get('address', '').strip()
        con.billing_consignor_type = req.POST.get('type', '').strip()
        con.billing_payment = req.POST.get('payment','').strip()
        con.billing_consignor_active = 'consignor_is_active' in req.POST
        
        con.save()

        return redirect('billing_shipper_manage')

    return render(req, 'accounts/billing_consignor_manage.html', {
        'shippers': shippers,
        'shipper': instance,
        'active_tab': mode,
        'is_edit': bool(billing_consignor_id),
        'search_query': query, 
    })

def billing_consignor_excel(req):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Billing Shipper"
    headers = [
        "Shipper Code", "Shipper Name", "Phone", "GST No",
        "GST Type", "Address", "Type","Active"
    ]
    sheet.append(headers)

    shippers = BillingConsignor.objects.all()

    for shipper in shippers:

        sheet.append([
            shipper.billing_consignor_code or "",
            shipper.billing_consignor_name,
            shipper.billing_consignor_phone,
            shipper.billing_consignor_gst,
            shipper.billing_consignor_gsttype,
            shipper.billing_consingor_address,
            shipper.billing_consignor_type,
            "YES" if shipper.billing_consignor_active else "NO",
        ])

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=billing_shippers.xlsx'

    workbook.save(response)
    return response

def billing_consignor_delete(req,billing_consignor_id):
    con = get_object_or_404(BillingConsignor,billing_consignor_id=billing_consignor_id)
    con.billing_consignor_active = False
    con.save()
    return redirect('shipper_manage')

def billing_consignee_manage(req,billing_consignee_id=None):
    query = req.GET.get('q', '').strip()

    receiver = BillingConsignee.objects.filter(billing_consignee_active=True).order_by('-billing_consignee_id')

    if query:
        receiver = receiver.filter(
            Q(billing_consignee_name__icontains=query) |
            Q(billing_consignee_phone__icontains=query) 
        )
    instance = None
    mode = req.GET.get('mode', 'list')
    if billing_consignee_id:
        mode = 'edit'
        instance = get_object_or_404(BillingConsignee, billing_consignee_id=billing_consignee_id)
    if req.method == 'POST':
        con = instance if billing_consignee_id else BillingConsignee()

        con.billing_consignee_name = req.POST.get('consignee_name', '').strip()
        con.billing_consignee_phone = req.POST.get('consignee_phone', '').strip()
        con.billing_consignee_address = req.POST.get('address', '').strip()
        con.billing_consignee_active = 'consignee_is_active' in req.POST

        con.save()

        return redirect('billing_receiver_manage')

    return render(req, 'accounts/billing_consignee_manage.html', {
        'receivers': receiver,
        'receiver': instance,
        'active_tab': mode,
        'is_edit': bool(billing_consignee_id),
        'search_query': query, 
    })

def billing_consignee_delete(req,billing_consignee_id):
    con = get_object_or_404(BillingConsignee,billing_consignee_id=billing_consignee_id)
    con.billing_consignee_active = False
    con.save()
    return redirect('billing_receiver_manage')

def courier_manage(req,id=None):
    query = req.GET.get('q', '').strip()

    couriers = CourierModel.objects.filter(courier_active=True).order_by('-id')

    if query:
        couriers = couriers.filter(courier_name__icontains=query)
    instance = None
    mode = req.GET.get('mode', 'list')
    if id:
        mode = 'edit'
        instance = get_object_or_404(CourierModel, id=id)
    if req.method == 'POST':
        con = instance if id else CourierModel()

        con.courier_name = req.POST.get('courier_name', '').strip()
        con.courier_active = 'courier_is_active' in req.POST
        con.save()

        return redirect('courier_manage')

    return render(req, 'accounts/courier_manage.html', {
        'couriers': couriers,
        'courier': instance,
        'active_tab': mode,
        'is_edit': bool(id),
        'search_query': query, 
    })

def courier_delete(req,id):
    con = get_object_or_404(CourierModel,id=id)
    con.courier_active = False
    con.save()
    return redirect('courier_manage')


@login_required
def create_billing(request):
    from_date = request.GET.get("from_date")
    to_date = request.GET.get("to_date")
    branch = request.user.branch
    consignors = BillingConsignor.objects.all()
    consignees = BillingConsignee.objects.all()
    search = request.GET.get("search")
    shipper = Consignor.objects.all().order_by('consignor_name')
    selected_consignors = request.GET.getlist("consignors")

    billed_cnotes = CnoteBilling.objects.filter(
        cnote=OuterRef('pk')
    )

    if request.user.is_superuser:
        cnotes = CnoteModel.objects.all()
    else:
        branch = request.user.branch
        cnotes = CnoteModel.objects.filter(
            booking_branch=branch
        )
    if search:
        numbers = re.split(r"[,\s]+", search.strip())
        numbers = [x for x in numbers if x]
        cnotes = cnotes.filter(cnote_number__in=numbers)

    if selected_consignors:
        cnotes = cnotes.filter(
            consignor__consignor_id__in=selected_consignors
        )
    if from_date and to_date:
        cnotes = cnotes.filter(date__range=[from_date, to_date])

    elif from_date:
        cnotes = cnotes.filter(date__gte=from_date)

    elif to_date:
        cnotes = cnotes.filter(date__lte=to_date)
    if request.method == "POST":
        consignor_id = request.POST.get("consignor")
        consignor = BillingConsignor.objects.get(billing_consignor_id=consignor_id)

        selected_ids = request.POST.getlist("selected")
        for cnote_id in selected_ids:

            inv_freight = request.POST.get(f"subtotal_{cnote_id}") or 0
            inv_lr = request.POST.get(f"inv_lr_{cnote_id}") or 0
            inv_unloading = request.POST.get(f"inv_unloading_{cnote_id}") or 0
            inv_other = request.POST.get(f"inv_other_{cnote_id}") or 0
            weight = request.POST.get(f"charged_weight_{cnote_id}")
            particulars = request.POST.get(f"particulars_{cnote_id}")
            cnote = CnoteModel.objects.get(cnote_id=cnote_id)

            CnoteBilling.objects.create(
                cnote=cnote,
                consignor=consignor,
                inv_freight=float(inv_freight),
                inv_lr=float(inv_lr),
                inv_unloading=float(inv_unloading),
                inv_other=float(inv_other),
                weight=weight,
                particulars=particulars,
                status=CnoteBilling.INV_PENDING
            )
        return redirect("view_cnote")

    return render(request, "accounts/create_billing.html", {
        "cnotes": cnotes,
        "consignors": consignors,
        "consignees":consignees,
        "selected_consignors": selected_consignors,
        "shippers":shipper
    })

@login_required
def view_cnote(req):
    if req.user.is_superuser:
        billing = CnoteBilling.objects.all()
        shipments = CourierShipment.objects.all()
    else:
        user_branch = req.user.branch

        billing = CnoteBilling.objects.filter(
            cnote__booking_branch=user_branch
        )

        shipments = CourierShipment.objects.filter(
            branch=user_branch
        )
    search_query = req.GET.get("search")
    if search_query:
        billing = billing.filter(
            Q(cnote__cnote_number__icontains=search_query) |
            Q(cnote__consignee__consignee_name__icontains=search_query)
        )

        shipments = shipments.filter(
            Q(cnote_number__icontains=search_query) |
            Q(consignee__billing_consignee_name__icontains=search_query)
        )
    billing_data = []
    for b in billing:
        total_item = b.cnote.total_item or 0
        inv_freight = b.inv_freight or 0
        inv_unloading = b.inv_unloading or 0
        inv_other = b.inv_other or 0
        inv_lr = b.inv_lr or 0

        rate = int(inv_freight / total_item) if total_item > 0 else 0
        other_charges = inv_unloading + inv_other
        total_amount = inv_freight + inv_lr + other_charges

        billing_data.append({
            "obj": b,
            "rate": rate,
            "consignee": b.cnote.consignee.consignee_name,
            "other": other_charges,
            "total_amount": total_amount,
            "type": "billing"
        })
    for s in shipments:

        freight = s.freight or 0
        lr = s.lr_charge or 0
        unloading = s.unloading_charge or 0
        other = s.other_charge or 0
        qty = s.qty or 0

        rate = round(freight / qty, 2) if qty > 0 else 0
        other_charges = unloading + other
        total_amount = freight + lr + other_charges

        billing_data.append({
            "obj": s, 
            "rate": rate,
            "consignee": s.consignee.billing_consignee_name if s.consignee else "",
            "other": other_charges,
            "total_amount": total_amount,
            "type": "shipment"
        })
    paginator = Paginator(billing_data, 10)  
    page_number = req.GET.get("page")
    page_obj = paginator.get_page(page_number)

    context = {
        "page_obj": page_obj,
        "search_query": search_query
    }
    return render(req,"accounts/view_cnote.html",context)


def export_cnote_excel(request):
    billing = CnoteBilling.objects.select_related(
        "cnote", "consignor", "cnote__consignee", "cnote__destination"
    )
    shipments = CourierShipment.objects.select_related(
        "consignor",
        "consignee"
    ).filter(status=CourierShipment.INV_PENDING)
    wb = Workbook()
    ws = wb.active
    ws.title = "Cnote Billing Report"

    headers = [
        "Date",
        "LR No",
        "Customer Code",
        "Invoice No",
        "Customer",
        "Consignee",
        "Destination",
        "Qty",
        "Payment",
        "Particulars",
        "Weight",
        "Rate",
        "Freight",
        "LR Charges",
        "Other Charges",
        "Billed Amount",
        "Booked Amount",
        "Status"
    ]

    ws.append(headers)

    for col in range(1, len(headers) + 1):
        ws.cell(row=1, column=col).font = Font(bold=True)

    row_num = 2

    for b in billing:
        total_item = b.cnote.total_item or 0
        inv_freight = b.inv_freight or 0
        inv_unloading = b.inv_unloading or 0
        inv_other = b.inv_other or 0
        inv_lr = b.inv_lr or 0

        rate = round(inv_freight / total_item, 2) if total_item > 0 else 0
        other_charges = inv_unloading + inv_other
        total_amount = inv_freight + inv_lr + other_charges

        ws.append([
            b.cnote.date.strftime("%d-%m-%Y"),
            b.cnote.cnote_number,
            b.consignor.billing_consignor_code,
            b.cnote.invoice_no,
            b.consignor.billing_consignor_name,  
            b.cnote.consignee.consignee_name,
            b.cnote.destination.location_name,
            b.cnote.total_item,
            b.consignor.billing_payment,
            b.particulars,
            b.weight,
            rate,
            inv_freight,
            inv_lr,
            other_charges,
            total_amount,
            b.cnote.total,
            b.status,
        ])

        row_num += 1
    for s in shipments:

        freight = s.freight or 0
        lr = s.lr_charge or 0
        unloading = s.unloading_charge or 0
        other = s.other_charge or 0
        qty = s.qty or 0

        rate = round(freight / qty, 2) if qty > 0 else 0
        other_charges = unloading + other
        total_amount = freight + lr + other_charges

        ws.append([
            s.booking_date.strftime("%d-%m-%Y") if s.booking_date else "",
            s.cnote_number,
            s.consignor.billing_consignor_code if s.consignor else "",
            "-", 
            s.consignor.billing_consignor_name if s.consignor else "",
            s.consignee.billing_consignee_name if s.consignee else "",
            s.destination,
            s.qty,
            s.consignor.billing_payment, 
            s.particulars,
            s.charged_weight,
            rate,
            freight,
            lr,
            other_charges,
            total_amount,
            total_amount,
            s.status,
        ])
        row_num += 1
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="Cnote_Billing_Report.xlsx"'

    wb.save(response)
    return response

def allocate_cnote(request):

    if request.method == "POST":
        branch_id = request.POST.get("branch")
        courier_id = request.POST.get("courier")
        cnote_numbers = request.POST.get("cnote_numbers")

        if branch_id and courier_id and cnote_numbers:

            branch = Branch.objects.get(branch_id=branch_id)
            courier = CourierModel.objects.get(id=courier_id)

            cnotes = cnote_numbers.splitlines()

            for cnote in cnotes:
                cnote = cnote.strip()

                if cnote:
                    CourierShipment.objects.create(
                        branch=branch,
                        courier=courier,
                        cnote_number=cnote,
                        qty=0,
                        rate=0,
                        actual_weight=0,
                        charged_weight=0,
                        freight=0,
                        lr_charge=0,
                        unloading_charge=0,
                        other_charge=0,
                        status=CourierShipment.ALLOCATE
                    )

            messages.success(request, "CNotes Allocated Successfully")
            return redirect("allocate_cnote")

    branches = Branch.objects.filter(branch_type="COMPANY")
    couriers = CourierModel.objects.all()
    allocated = CourierShipment.objects.filter(status=CourierShipment.ALLOCATE)

    return render(request, "accounts/allocate_cnote.html", {
        "branches": branches,
        "couriers": couriers,
        "allocated": allocated
    })

def edit_shipment(request, id):

    shipment = get_object_or_404(CourierShipment, pk=id)
    consignors = BillingConsignor.objects.all()
    consignees = BillingConsignee.objects.all()
    if request.method == "POST":
        shipment.consignor_id = request.POST.get('consignor')
        shipment.consignee_id = request.POST.get('consignee')
        shipment.booking_date = request.POST.get("booking_date")
        shipment.destination = request.POST.get("destination").upper()
        shipment.particulars = request.POST.get("particulars")
        shipment.qty = request.POST.get("qty")
        shipment.rate = request.POST.get("rate")
        shipment.actual_weight = request.POST.get("actual_weight")
        shipment.charged_weight = request.POST.get("charged_weight")
        shipment.freight = request.POST.get("freight")
        shipment.lr_charge = request.POST.get("lr_charge")
        shipment.unloading_charge = request.POST.get("unloading_charge")
        shipment.other_charge = request.POST.get("other_charge")
        shipment.status = CourierShipment.INV_PENDING

        shipment.save()

        return redirect("allocate_cnote")

    return render(request, "accounts/edit_shipment.html", {
        "shipment": shipment,
        "consignors": consignors,
        "consignees": consignees
    })


def create_invoice(request):

    credit_customers = BillingConsignor.objects.filter(
        billing_payment='CREDIT'
    )

    customer_id = request.GET.get('customer')
    from_date = request.GET.get('from_date')
    to_date = request.GET.get('to_date')

    searched = False
    cnote_list = []
    selected_customer_name = ""

    if customer_id and from_date and to_date:
        searched = True

        shipments = CourierShipment.objects.filter(
            status=CourierShipment.INV_PENDING,
            consignor__billing_payment='CREDIT',
            consignor__billing_consignor_id=customer_id,
            booking_date__range=[from_date, to_date]
        )

        billings = CnoteBilling.objects.filter(
            status=CnoteBilling.INV_PENDING,
            consignor__billing_payment='CREDIT',
            consignor__billing_consignor_id=customer_id,
            cnote__date__range=[from_date, to_date]
        )

        for s in shipments:
            total_amount = (
                (s.freight or 0) +
                (s.lr_charge or 0) +
                (s.unloading_charge or 0) +
                (s.other_charge or 0)
            )

            cnote_list.append({
                'id': s.id,
                'cnote_number': s.cnote_number,
                'consignor': s.consignor.billing_consignor_name,
                'consignee': s.consignee.billing_consignee_name if s.consignee else '',
                'qty': s.qty or 0,
                'total_amount': total_amount,
                'type': 'COURIER'
            })

        for b in billings:
            total_amount = (
                (b.inv_freight or 0) +
                (b.inv_lr or 0) +
                (b.inv_unloading or 0) +
                (b.inv_other or 0)
            )

            cnote_list.append({
                'id': b.id,
                'cnote_number': b.cnote.cnote_number,
                'consignor': b.consignor.billing_consignor_name,
                'consignee': b.cnote.consignee.consignee_name if b.cnote.consignee else '',
                'qty': b.cnote.total_item or 0,
                'total_amount': total_amount,
                'type': 'PARCEL'
            })

        cnote_list = sorted(cnote_list, key=lambda x: x['cnote_number'])

        selected = BillingConsignor.objects.filter(billing_consignor_id=customer_id).first()
        if selected:
            selected_customer_name = selected.billing_consignor_name.upper()

    if request.method == "POST":

        invoice_number = request.POST.get("invoice_number")
        invoice_date = request.POST.get("invoice_date")
        customer_id = request.POST.get("customer")
        selected_cnotes = request.POST.getlist("cnotes")

        if invoice_number and invoice_date and customer_id and selected_cnotes:
            customer = BillingConsignor.objects.get(billing_consignor_id=customer_id)

            invoice = Invoice.objects.create(
                invoice_number=invoice_number,
                invoice_date=invoice_date,
                customer_id=customer_id,
                invoiced_by=request.user.username
            )

            for item in selected_cnotes:
                obj_id, obj_type = item.split("|")

                if obj_type == "COURIER":
                    shipment = CourierShipment.objects.get(id=obj_id)
                    shipment.invoice = invoice
                    shipment.status = CourierShipment.INV_INVOICED
                    shipment.save()

                elif obj_type == "PARCEL":
                    billing = CnoteBilling.objects.get(id=obj_id)
                    billing.invoice = invoice
                    billing.status = CnoteBilling.INV_INVOICED
                    billing.save()

            total = invoice.total_amount()
            invoice.save()
            invoice.create_accounting_entry()
            return redirect("view_invoice")

    context = {
        'customers': credit_customers,
        'cnotes': cnote_list,
        'searched': searched,
        'selected_customer': customer_id,
        'selected_customer_name': selected_customer_name,
        'from_date': from_date,
        'to_date': to_date,
        'today': timezone.now().date(),
    }

    return render(request, "accounts/create_invoice.html", context)

def invoice_list(request):

    invoices = Invoice.objects.select_related('customer').all().order_by('-invoice_date')

    invoice_data = []
    for inv in invoices:
        type = inv.customer.billing_consignor_gsttype
        cgst = inv.cgst or 0
        sgst = inv.sgst or 0
        igst = inv.igst or 0
        round_off = inv.roundoff or 0
        total = inv.total_amount()
        taxable_value = total - (cgst + sgst + igst)
        if round_off<=0.5:
            total = total+round_off
        else:
            total = total-round_off
        if type == "REVERSE":
            taxable_value = total
        invoice_data.append({
            'id': inv.id,
            'invoice_number': inv.invoice_number,
            'invoice_date': inv.invoice_date,
            'customer': inv.customer.billing_consignor_name,
            'type': type,
            'cgst': cgst,
            'sgst': sgst,
            'igst': igst,
            'roundoff':round_off,
            'taxable_value': taxable_value,
            'total_amount': total,
        })

    context = {
        'invoices': invoice_data,
        'today': timezone.now().date()
    }

    return render(request, "accounts/invoice_list.html", context)